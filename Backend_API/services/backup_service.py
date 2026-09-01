"""Backup / restore service (JSON files, optional ZIP with images, Excel XLSX)."""
import json
import os
import re
import shutil
import zipfile
from datetime import datetime

from sqlalchemy import func, or_

import models
from config import config
from database import SessionLocal

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    OPENPYXL_AVAILABLE = False


def _json_safe(value):
    """Serialize lists/dicts to JSON strings so they fit in an Excel cell."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    try:
        return json.dumps(value, ensure_ascii=False)
    except Exception:
        return str(value)


def _parse_cell(value):
    """Reverse of _json_safe: parse JSON-looking strings back into objects."""
    if isinstance(value, str):
        s = value.strip()
        if s.startswith("[") or s.startswith("{"):
            try:
                return json.loads(s)
            except Exception:
                return value
    return value


def _style_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E8C")
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _write_sheet(wb, name, rows):
    ws = wb.create_sheet(name)
    if not rows:
        ws.append(["no_data"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([_json_safe(row.get(h)) for h in headers])
    _style_header(ws)
    for i, h in enumerate(headers, start=1):
        letter = openpyxl.utils.get_column_letter(i) if hasattr(openpyxl, "utils") else chr(64 + min(i, 26))
        ws.column_dimensions[letter].width = min(42, max(10, len(str(h)) + 4))



def _collect_shop_data(db, shop_id: int) -> dict:
    shop = db.query(models.Shop).filter(models.Shop.id == shop_id).first()
    if not shop:
        return None
    return {
        "version": 1,
        "exported_at": datetime.utcnow().isoformat(),
        "type": "shop",
        "shop": shop.to_dict(include_private=True),
        "categories": [c.to_dict() for c in db.query(models.Category).filter(models.Category.shop_id == shop_id).all()],
        "products": [p.to_dict() for p in db.query(models.Product).filter(models.Product.shop_id == shop_id).all()],
        "orders": [o.to_dict() for o in db.query(models.Order).filter(models.Order.shop_id == shop_id).all()],
        "customers": [c.to_dict() for c in db.query(models.Customer).filter(models.Customer.shop_id == shop_id).all()],
        "settings": [s.to_dict() for s in db.query(models.Setting).filter(models.Setting.shop_id == shop_id).all()],
    }


def _collect_system_data(db) -> dict:
    return {
        "version": 1,
        "exported_at": datetime.utcnow().isoformat(),
        "type": "system",
        "shops": [s.to_dict(include_private=True) for s in db.query(models.Shop).all()],
        "users": [{
            "id": u.id, "username": u.username, "email": u.email,
            "password_hash": u.password_hash, "role": u.role,
            "shop_id": u.shop_id, "status": u.status,
        } for u in db.query(models.User).all()],
        "categories": [c.to_dict() for c in db.query(models.Category).all()],
        "products": [p.to_dict() for p in db.query(models.Product).all()],
        "orders": [o.to_dict() for o in db.query(models.Order).all()],
        "customers": [c.to_dict() for c in db.query(models.Customer).all()],
        "settings": [s.to_dict() for s in db.query(models.Setting).all()],
    }


def _walk_strings(obj):
    """Yield every string value inside nested dicts / lists."""
    if isinstance(obj, dict):
        for v in obj.values():
            yield from _walk_strings(v)
    elif isinstance(obj, list):
        for v in obj:
            yield from _walk_strings(v)
    elif isinstance(obj, str):
        yield obj


def _collect_refs(data, prefix):
    """Unique string values in the backup data that start with a prefix."""
    return sorted({s for s in _walk_strings(data) if s.startswith(prefix)})


def _rewrite_upload_refs(data, mapping):
    """Deep-copy the backup data, replacing exact string matches via mapping."""
    if isinstance(data, dict):
        return {k: _rewrite_upload_refs(v, mapping) for k, v in data.items()}
    if isinstance(data, list):
        return [_rewrite_upload_refs(v, mapping) for v in data]
    if isinstance(data, str):
        return mapping.get(data, data)
    return data


def _restore_images(data, image_root):
    """Copy bundled images back into UPLOAD_DIR and rewrite refs to /uploads/.

    Backup ZIPs embed the real image files under an "images/" folder and the
    JSON stores references like "images/products/a.jpg". On import we extract
    them to config.UPLOAD_DIR and rewrite the references to the public
    "/uploads/products/a.jpg" form so the storefront can display them.

    Returns (data, images_restored).
    """
    if not image_root or not os.path.isdir(image_root):
        return data, 0
    mapping = {}
    count = 0
    for ref in _collect_refs(data, "images/"):
        rel = ref[len("images/"):].lstrip("/")
        if not rel:
            continue
        src = os.path.join(image_root, rel.replace("/", os.sep))
        if not os.path.isfile(src):
            continue
        dest = os.path.join(config.UPLOAD_DIR, rel.replace("/", os.sep))
        try:
            os.makedirs(os.path.dirname(dest), exist_ok=True)
            shutil.copyfile(src, dest)
        except Exception:
            continue
        mapping[ref] = "/uploads/" + rel.replace(os.sep, "/")
        count += 1
    if mapping:
        data = _rewrite_upload_refs(data, mapping)
    return data, count


def create_backup(db, shop_id=None, fmt="") -> dict:
    """Create a JSON backup file, optionally wrapped in a ZIP. Returns file info.

    fmt: "" (auto) -> JSON unless the file is large (>200KB), then ZIP.
         "zip"     -> always produce a .zip.
         "json"    -> always produce a .json.
    """
    os.makedirs(config.BACKUP_DIR, exist_ok=True)
    if shop_id:
        data = _collect_shop_data(db, shop_id)
        kind = "shop"
        prefix = f"shop_{shop_id}"
    else:
        data = _collect_system_data(db)
        kind = "system"
        prefix = "system"

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"{prefix}_{ts}.json"
    filepath = os.path.join(config.BACKUP_DIR, filename)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # Wrap into ZIP when requested or when the file is large. ZIP backups also
    # embed the real image files (product photos, logos, banners, category
    # images, slideshow, QR codes, receipts) under an "images/" folder and the
    # JSON references are rewritten to "images/<path>" so the bundle is fully
    # portable to another workspace.
    fmt = (fmt or "").lower()
    should_zip = fmt == "zip" or (fmt != "json" and os.path.getsize(filepath) > 200 * 1024)
    if should_zip:
        mapping = {}
        for ref in _collect_refs(data, "/uploads/"):
            rel = ref[len("/uploads/"):].lstrip("/")
            src = os.path.join(config.UPLOAD_DIR, rel.replace("/", os.sep))
            if rel and os.path.isfile(src):
                mapping[ref] = "images/" + rel.replace(os.sep, "/")
        if mapping:
            data = _rewrite_upload_refs(data, mapping)
        zip_name = filename.replace(".json", ".zip")
        zip_path = os.path.join(config.BACKUP_DIR, zip_name)
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for ref, newref in mapping.items():
                rel = newref[len("images/"):]
                zf.write(os.path.join(config.UPLOAD_DIR, rel.replace("/", os.sep)),
                         arcname=newref)
            zf.writestr(filename, json.dumps(data, ensure_ascii=False, indent=2))
        os.remove(filepath)
        filepath = zip_path
        filename = zip_name

    history = models.BackupHistory(shop_id=shop_id, filename=filename, filepath=filepath, kind=kind)
    db.add(history)
    db.commit()
    db.refresh(history)

    return {"id": history.id, "filename": filename, "filepath": filepath.replace("\\", "/"),
            "kind": kind, "created_at": history.created_at.isoformat(), "size": os.path.getsize(filepath)}


def export_backup_excel(db, shop_id=None) -> str:
    """Export shop/system data as an .xlsx workbook. Returns the file path."""
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is not installed; cannot export Excel")

    os.makedirs(config.BACKUP_DIR, exist_ok=True)
    if shop_id:
        data = _collect_shop_data(db, shop_id)
        prefix = f"shop_{shop_id}"
        kind = "shop"
    else:
        data = _collect_system_data(db)
        prefix = "system"
        kind = "system"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "overview"
    ws.append(["key", "value"])
    ws.append(["type", data.get("type", "")])
    ws.append(["version", data.get("version", 1)])
    ws.append(["exported_at", data.get("exported_at", "")])
    _style_header(ws)

    if data.get("shops"):
        _write_sheet(wb, "shops", data["shops"])
    if data.get("users"):
        _write_sheet(wb, "users", data["users"])
    if data.get("shop"):
        _write_sheet(wb, "shop", [data["shop"]])
    _write_sheet(wb, "categories", data.get("categories", []))
    _write_sheet(wb, "products", data.get("products", []))
    _write_sheet(wb, "orders", data.get("orders", []))
    _write_sheet(wb, "customers", data.get("customers", []))
    _write_sheet(wb, "settings", data.get("settings", []))

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"{prefix}_{ts}.xlsx"
    filepath = os.path.join(config.BACKUP_DIR, filename)
    wb.save(filepath)

    history = models.BackupHistory(shop_id=shop_id, filename=filename, filepath=filepath, kind=kind)
    db.add(history)
    db.commit()
    return filepath


def _xlsx_to_dict(filepath: str) -> dict:
    """Read an exported .xlsx workbook back into the backup data structure."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    def read_sheet(name):
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() for h in rows[0] if h is not None and str(h).strip() != ""]
        out = []
        for r in rows[1:]:
            if not any(c is not None and str(c).strip() not in ("", "no_data") for c in r):
                continue
            d = {}
            for i, h in enumerate(headers):
                if i < len(r):
                    d[h] = _parse_cell(r[i])
            out.append(d)
        return out

    data = {"type": "system", "categories": [], "products": [], "orders": [],
            "customers": [], "settings": []}
    shops = read_sheet("shops")
    if shops:
        data["shops"] = shops
    users = read_sheet("users")
    if users:
        data["users"] = users
    shop_rows = read_sheet("shop")
    if shop_rows:
        data["shop"] = shop_rows[0]
    data["categories"] = read_sheet("categories")
    data["products"] = read_sheet("products")
    data["orders"] = read_sheet("orders")
    data["customers"] = read_sheet("customers")
    data["settings"] = read_sheet("settings")

    if "overview" in wb.sheetnames:
        for r in wb["overview"].iter_rows(values_only=True):
            if r and len(r) > 1 and str(r[0]).strip() == "type":
                data["type"] = str(r[1]).strip()
    return data



def _read_backup(filepath: str):
    """Load a backup file and return (data, image_root).

    image_root is the extracted "images/" folder for ZIP backups that embed
    image files (or None when the backup has no image bundle). The JSON data
    returned still uses portable "images/<path>" references — call
    _restore_images() to materialize them into the uploads directory.
    """
    if filepath.endswith(".xlsx"):
        return _xlsx_to_dict(filepath), None
    if filepath.endswith(".zip"):
        extract_dir = os.path.join(config.BACKUP_DIR, "_extract")
        shutil.rmtree(extract_dir, ignore_errors=True)
        os.makedirs(extract_dir, exist_ok=True)
        json_name = None
        with zipfile.ZipFile(filepath, "r") as zf:
            for name in zf.namelist():
                if name.endswith(".json") and not name.startswith("images/"):
                    json_name = name
                    break
            if json_name is None:
                raise ValueError("ZIP backup contains no JSON data file")
            zf.extractall(extract_dir)
        json_path = os.path.join(extract_dir, json_name)
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        image_root = os.path.join(extract_dir, "images")
        if not os.path.isdir(image_root):
            image_root = None
        return data, image_root
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f), None


def _order_core(number: str) -> str:
    """Return the base order number with any import suffixes stripped.

    Real order numbers end with a 14-digit timestamp (yyyymmddHHMMSS); imports
    append extra -HHMMSS (6-digit) groups, possibly more than once.
    """
    core = number or ""
    while re.search(r"-\d{6}$", core) and not re.search(r"-\d{14}$", core):
        core = re.sub(r"-\d{6}$", "", core)
    return core


def _restore_shop_contents(db, shop_id: int, data: dict) -> tuple:
    """Restore categories / products / customers / orders / settings for one shop.

    Categories and products are matched by name (case-insensitive). Existing
    ones are skipped instead of duplicated. Returns (added, skipped).
    """
    added = 0
    skipped = 0
    id_map = {}
    seen_cat_ids = {}
    seen_products = set()

    # Restore categories (dedup by name, map old ids -> new ids)
    for c in data.get("categories", []):
        old_id = c.get("id")
        name = (c.get("name") or "").strip()
        key = name.lower()
        if key and key in seen_cat_ids:
            id_map[old_id] = seen_cat_ids[key]
            skipped += 1
            continue
        if key:
            existing = db.query(models.Category).filter(
                models.Category.shop_id == shop_id,
                func.lower(models.Category.name) == key).first()
            if existing:
                id_map[old_id] = existing.id
                seen_cat_ids[key] = existing.id
                skipped += 1
                continue
        cat = models.Category(
            shop_id=shop_id, name=c.get("name", ""), slug=c.get("slug", ""),
            parent_id=c.get("parent_id"), image=c.get("image", ""),
            sort_order=c.get("sort_order", 0))
        db.add(cat)
        db.flush()
        id_map[old_id] = cat.id
        if key:
            seen_cat_ids[key] = cat.id
        added += 1

    # Restore products (dedup by name — do not add the same product twice)
    for p in data.get("products", []):
        name = (p.get("name") or "").strip()
        key = name.lower()
        if key and key in seen_products:
            skipped += 1
            continue
        if key:
            existing = db.query(models.Product).filter(
                models.Product.shop_id == shop_id,
                func.lower(models.Product.name) == key).first()
            if existing:
                seen_products.add(key)
                skipped += 1
                continue
        old_cat = p.get("category_id")
        db.add(models.Product(
            shop_id=shop_id,
            category_id=id_map.get(old_cat, None),
            name=p.get("name", ""),
            description=p.get("description", ""),
            price=p.get("price", 0),
            sale_price=p.get("sale_price"),
            quantity=p.get("quantity", 0),
            images=models.JSONText.dumps(p.get("images", [])),
            custom_attributes=models.JSONText.dumps(p.get("custom_attributes", [])),
            variations=models.JSONText.dumps(p.get("variations", [])),
            metadata_json=models.JSONText.dumps(p.get("metadata", {})),
            featured=p.get("featured", False),
            status=p.get("status", "active"),
        ))
        if key:
            seen_products.add(key)
        added += 1

    # Restore customers (dedup by phone or telegram)
    seen_cust = set()
    for c in data.get("customers", []):
        phone = (c.get("phone") or "").strip()
        telegram = (c.get("telegram") or "").strip()
        key = phone or telegram
        if key and key in seen_cust:
            skipped += 1
            continue
        if key:
            ors = []
            if phone:
                ors.append(models.Customer.phone == phone)
            if telegram:
                ors.append(models.Customer.telegram == telegram)
            existing = db.query(models.Customer).filter(
                models.Customer.shop_id == shop_id, or_(*ors)).first() if ors else None
            if existing:
                seen_cust.add(key)
                skipped += 1
                continue
        db.add(models.Customer(
            shop_id=shop_id, name=c.get("name", ""), phone=c.get("phone", ""),
            telegram=c.get("telegram", ""), email=c.get("email", ""),
            address=c.get("address", ""), city=c.get("city", ""),
            country=c.get("country", ""), notes=c.get("notes", "")))
        if key:
            seen_cust.add(key)
        added += 1

    # Restore orders (dedup by the normalized base order number)
    suffix = datetime.utcnow().strftime("%H%M%S")
    existing_cores = {_order_core(o.order_number) for o in
                      db.query(models.Order).filter(models.Order.shop_id == shop_id).all()}
    seen_orders = set()
    for o in data.get("orders", []):
        number = (o.get("order_number") or "").strip()
        core = _order_core(number)
        if core and (core in seen_orders or core in existing_cores):
            skipped += 1
            continue
        order = models.Order(
            shop_id=shop_id,
            order_number=f"{core}-{suffix}" if core else f"IMP-{suffix}",
            customer_name=o.get("customer_name", ""),
            customer_email=o.get("customer_email", ""),
            customer_phone=o.get("customer_phone", ""),
            customer_telegram=o.get("customer_telegram", ""),
            customer_address=o.get("customer_address", ""),
            customer_city=o.get("customer_city", ""),
            customer_country=o.get("customer_country", ""),
            customer_note=o.get("customer_note", ""),
            items_total=o.get("items_total", 0),
            shipping_fee=o.get("shipping_fee", 0),
            discount=o.get("discount", 0),
            total=o.get("total", 0),
            currency=o.get("currency", "USD"),
            payment_status=o.get("payment_status", "pending"),
            order_status=o.get("order_status", "pending"),
        )
        db.add(order)
        db.flush()
        for it in o.get("items", []):
            db.add(models.OrderItem(
                order_id=order.id, product_id=it.get("product_id"),
                product_name=it.get("product_name", ""), price=it.get("price", 0),
                quantity=it.get("quantity", 1),
                variations=models.JSONText.dumps(it.get("variations", {}))))
        if core:
            seen_orders.add(core)
            existing_cores.add(core)
        added += 1 + len(o.get("items", []))

    # Restore settings (dedup by key)
    seen_settings = set()
    for s in data.get("settings", []):
        k = (s.get("key") or "").strip()
        if k in seen_settings:
            skipped += 1
            continue
        if k:
            existing = db.query(models.Setting).filter(
                models.Setting.shop_id == shop_id,
                models.Setting.key == k).first()
            if existing:
                seen_settings.add(k)
                skipped += 1
                continue
        db.add(models.Setting(shop_id=shop_id, key=k, value=s.get("value", "")))
        if k:
            seen_settings.add(k)
        added += 1

    return added, skipped


def import_shop_backup(db, shop_id: int, filepath: str):
    """Import a shop backup into an existing shop.

    Returns (records_restored, duplicates_skipped, images_restored).
    """
    data, image_root = _read_backup(filepath)
    data, images_restored = _restore_images(data, image_root)
    if data.get("type") == "system":
        raise ValueError("This is a system backup, not a shop backup")

    count = 0
    shop_data = data.get("shop") or {}

    # Update shop profile (keep aba/telegram settings from backup)
    shop = db.query(models.Shop).filter(models.Shop.id == shop_id).first()
    if shop and shop_data:
        shop.shop_name = shop_data.get("shop_name", shop.shop_name)
        shop.logo = shop_data.get("logo", shop.logo)
        shop.banner = shop_data.get("banner", shop.banner)
        shop.bio = shop_data.get("bio", shop.bio)
        shop.description = shop_data.get("description", shop.description)
        shop.slideshow = models.JSONText.dumps(shop_data.get("slideshow", []))
        shop.social_media = models.JSONText.dumps(shop_data.get("social_media", {}))
        shop.theme = models.JSONText.dumps(shop_data.get("theme", {}))
        if shop_data.get("aba_settings") is not None:
            shop.aba_settings = models.JSONText.dumps(shop_data.get("aba_settings"))
        if shop_data.get("telegram_settings") is not None:
            shop.telegram_settings = models.JSONText.dumps(shop_data.get("telegram_settings"))
        shop.contact = shop_data.get("contact", shop.contact)
        db.commit()
        count += 1

    # Restore categories, products, customers, orders, settings
    added, skipped = _restore_shop_contents(db, shop_id, data)
    count += added

    db.commit()
    return count, skipped, images_restored


def import_system_backup(db, filepath: str):
    """Import a full system backup (shops, users + every shop's data).

    Returns (records_restored, duplicates_skipped, images_restored).
    """
    data, image_root = _read_backup(filepath)
    data, images_restored = _restore_images(data, image_root)
    count = 0

    # Restore shops (create missing ones), building old -> new id map
    shop_id_map = {}
    for s in data.get("shops", []):
        existing = db.query(models.Shop).filter(models.Shop.username == s["username"]).first()
        if existing:
            shop_id_map[s.get("id")] = existing.id
        else:
            new_shop = models.Shop(
                username=s["username"], shop_name=s.get("shop_name", ""),
                logo=s.get("logo", ""), banner=s.get("banner", ""),
                bio=s.get("bio", ""), description=s.get("description", ""),
                slideshow=models.JSONText.dumps(s.get("slideshow", [])),
                social_media=models.JSONText.dumps(s.get("social_media", {})),
                theme=models.JSONText.dumps(s.get("theme", {})),
                aba_settings=models.JSONText.dumps(s.get("aba_settings", {})),
                telegram_settings=models.JSONText.dumps(s.get("telegram_settings", {})),
                currency=s.get("currency", "USD"), status=s.get("status", "active"),
                contact=s.get("contact", ""))
            if s.get("expires_at"):
                try:
                    new_shop.expires_at = datetime.fromisoformat(str(s["expires_at"]).replace("Z", "+00:00")).replace(tzinfo=None)
                except Exception:
                    pass
            db.add(new_shop)
            db.flush()
            shop_id_map[s.get("id")] = new_shop.id
        count += 1

    # Restore users (remap shop references through the new shop ids)
    for u in data.get("users", []):
        new_shop_id = shop_id_map.get(u.get("shop_id"))
        if new_shop_id is None and u.get("shop_id") is not None:
            new_shop_id = u.get("shop_id")  # keep original reference when shop already existed
        existing = db.query(models.User).filter(models.User.username == u["username"]).first()
        if existing:
            existing.email = u.get("email", existing.email)
            existing.role = u.get("role", existing.role)
            existing.status = u.get("status", existing.status)
            if new_shop_id is not None:
                existing.shop_id = new_shop_id
        else:
            db.add(models.User(
                username=u["username"], email=u.get("email", ""),
                password_hash=u.get("password_hash", ""), role=u.get("role", "shop_owner"),
                shop_id=new_shop_id, status=u.get("status", "active")))
        count += 1

    db.flush()

    # Restore each shop's categories / products / customers / orders / settings
    from collections import defaultdict
    by_shop = defaultdict(lambda: {"categories": [], "products": [], "customers": [],
                                   "orders": [], "settings": []})
    for c in data.get("categories", []):
        by_shop[c.get("shop_id")]["categories"].append(c)
    for p in data.get("products", []):
        by_shop[p.get("shop_id")]["products"].append(p)
    for c in data.get("customers", []):
        by_shop[c.get("shop_id")]["customers"].append(c)
    for o in data.get("orders", []):
        by_shop[o.get("shop_id")]["orders"].append(o)
    for s in data.get("settings", []):
        by_shop[s.get("shop_id")]["settings"].append(s)

    for old_shop_id, contents in by_shop.items():
        new_shop_id = shop_id_map.get(old_shop_id) or old_shop_id
        added, skipped = _restore_shop_contents(db, new_shop_id, contents)
        count += added

    db.commit()
    return count, skipped, images_restored


