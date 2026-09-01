"""Plan / self-serve shop registration + reseller commission endpoints.

- Anyone can create their own shop from the storefront, pick a plan and pay
  via ABA (the platform merchant shop collects the plan payment).
- On confirmed payment the shop is activated with the plan's limits + expiry.
- Admins can create resellers; shops registered with a reseller's referral
  code earn that reseller a commission on the plan price.
"""
import io
import json
import os
import tempfile
import time
import zipfile

from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

import models
import schemas
from config import config as app_config
from database import get_db
from security import (create_access_token, get_current_admin, get_current_user,
                      hash_password, log_activity, require_shop_access)
from services import aba_service, backup_service
from services.aba_service import PaymentNotConfigured

router = APIRouter(prefix="/api/plans", tags=["plans"])

# Platform merchant that collects plan payments (must have ABA configured).
PLATFORM_SHOP_USERNAME = os.getenv("PLATFORM_SHOP_USERNAME", "demo")
# Public storefront domain used in referral links (frontend_user).
STORE_URL = os.getenv("STORE_URL", "http://localhost:3000")

# Starter plan is FREE for customers (1-month plan). The free shop-opening
# registration offer ends on 07/09/2026.
OFFER_ENDS = "2026-09-07"

PLANS = {
    "starter": {"name": "Starter", "price": 9.99, "days": 30,
                "max_products": 20, "max_categories": 10,
                "free": True, "offer_ends": OFFER_ENDS},
    "growth": {"name": "Growth", "price": 55.99, "days": 180,
               "max_products": 50, "max_categories": 20},
    "premium": {"name": "Premium", "price": 99.99, "days": 365,
                "max_products": 200, "max_categories": 30},
}


@router.get("")
def list_plans():
    """Public: available plans with price, period and limits."""
    return [{"id": k, **v} for k, v in PLANS.items()]


def _gen_order_number(prefix="PLN"):
    return f"{prefix}{int(time.time() * 1000)}"


def _find_reseller(db, code):
    ref = (code or "").strip().lower()
    if not ref:
        return None
    for r in db.query(models.User).filter(models.User.role == "reseller").all():
        if (r.referral_code or "").strip().lower() == ref:
            return r
    return None


@router.post("/register")
def register_shop_plan(data: schemas.ShopRegister, db: Session = Depends(get_db)):
    """Public: create your own shop + owner account, choose a plan, pay via ABA."""
    username = (data.username or "").strip().lower()
    if not username or not data.password:
        raise HTTPException(status_code=400, detail="Username and password are required")
    if db.query(models.Shop).filter(models.Shop.username == username).first():
        raise HTTPException(status_code=400, detail="Username already exists")
    if db.query(models.User).filter(models.User.username == username).first():
        raise HTTPException(status_code=400, detail="Username already exists")
    plan = PLANS.get((data.plan or "starter").strip().lower())
    if not plan:
        raise HTTPException(status_code=400, detail="Invalid plan")

    reseller = _find_reseller(db, data.referral_code)
    if (data.referral_code or "").strip() and not reseller:
        raise HTTPException(status_code=400, detail="Invalid referral code")

    # Reseller promo discount (0 to reseller.discount_max, default up to $1).
    # If the customer didn't ask for a discount, the reseller's saved promo
    # discount is applied automatically.
    discount = 0.0
    if reseller:
        requested = float(data.discount or 0)
        if requested <= 0:
            requested = float(getattr(reseller, "promo_discount", 0) or 0)
        discount = max(0.0, min(requested, float(reseller.discount_max or 1.0)))
    amount = round(max(0.01, plan["price"] - discount), 2)

    # The Starter plan is FREE (1 month) for customers: no payment needed — the
    # shop opens right away for the full 1 month at $0.
    free_plan = bool(plan.get("free"))
    if free_plan:
        amount = 0.0
        discount = 0.0

    # 1) Create the shop (pending until paid — unless the plan is free).
    shop = models.Shop(
        username=username,
        shop_name=data.shop_name or username,
        currency=data.currency or "USD",
        status="active" if free_plan else "pending",
        max_products=plan["max_products"],
        max_categories=plan["max_categories"],
        plan=(data.plan or "starter").strip().lower(),
        plan_price=amount,
        plan_discount=round(discount, 2),
        reseller_id=reseller.id if reseller else None,
        plan_started_at=datetime.utcnow() if free_plan else None,
        expires_at=(datetime.utcnow() + timedelta(days=plan["days"])) if free_plan else None,
    )
    db.add(shop)
    db.flush()

    # 2) Owner account.
    owner = models.User(
        username=username,
        email=data.email,
        password_hash=hash_password(data.password),
        role="shop_owner",
        shop_id=shop.id,
    )
    db.add(owner)

    # 3) Plan order on the platform merchant. Free plans are recorded as paid
    #    ($0) for audit; paid plans get an ABA (KHQR) payment.
    payment = None
    order_id = None
    platform = db.query(models.Shop).filter(models.Shop.username == PLATFORM_SHOP_USERNAME).first()
    if platform:
        order = models.Order(
            shop_id=platform.id,
            order_number=_gen_order_number(),
            customer_name=data.shop_name or username,
            customer_email=data.email,
            customer_phone=data.phone,
            items_total=amount,
            total=amount,
            currency=shop.currency,
            payment_method="free" if free_plan else "khqr",
            payment_status="paid" if free_plan else "pending",
            paid_at=datetime.utcnow() if free_plan else None,
            order_status="active" if free_plan else "pending",
        )
        db.add(order)
        db.flush()
        item_name = f"Plan {plan['name']} ({plan['days']} days)"
        if free_plan:
            item_name += " — FREE (1 month)"
        elif discount:
            item_name += f" (promo -${round(discount, 2)})"
        db.add(models.OrderItem(
            order_id=order.id,
            product_name=item_name,
            price=amount,
            quantity=1,
        ))
        if not free_plan:
            try:
                payment = aba_service.build_checkout_url(order, platform, success_url=data.success_url)
                order.transaction_id = payment.get("transaction_id", "")
            except PaymentNotConfigured:
                payment = None
        order_id = order.id

    db.commit()
    log_activity(db, "plan_register",
                 f"Shop {shop.username} registered with plan {plan['name']}"
                 + (f" (FREE 1 month)" if free_plan else "")
                 + (f" (reseller {reseller.username}, promo -${round(discount, 2)})" if reseller else ""),
                 shop.id)
    db.commit()

    # Auto-login: the new owner is immediately logged in on the storefront.
    owner_token = create_access_token({"sub": str(owner.id), "role": owner.role})

    return {
        "shop_id": shop.id,
        "username": username,
        "dashboard_url": os.getenv("DASHBOARD_URL", "http://localhost:3002/"),
        "plan": {"id": shop.plan, **plan},
        "amount": amount,
        "discount": round(discount, 2),
        "order_id": order_id,
        "payment": payment,
        "free": bool(free_plan),
        "offer_ends": plan.get("offer_ends"),
        "expires_at": shop.expires_at.isoformat() + "Z" if shop.expires_at else None,
        "access_token": owner_token,
        "token_type": "bearer",
        "user": {"id": owner.id, "username": owner.username, "email": owner.email,
                 "role": owner.role, "shop_id": owner.shop_id, "status": owner.status},
    }
@router.post("/confirm")
def confirm_plan_payment(data: schemas.PlanConfirm, db: Session = Depends(get_db)):
    """Verify the plan payment; on success activate/extend the shop's plan."""
    order = db.query(models.Order).filter(models.Order.id == data.order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    shop = db.query(models.Shop).filter(models.Shop.id == data.shop_id).first()
    if not shop:
        raise HTTPException(status_code=404, detail="Shop not found")
    platform = db.query(models.Shop).filter(models.Shop.id == order.shop_id).first()

    # Which plan was ordered? Shop upgrades store it on the order note.
    plan_id = shop.plan
    note = (order.customer_note or "")
    if note.startswith("PLAN_UPGRADE:"):
        plan_id = note.split(":", 1)[1].strip() or plan_id
    plan = PLANS.get(plan_id, {})

    # Free plan ($0, already recorded as paid) — nothing to verify.
    if order.payment_status == "paid" and float(order.total or 0) == 0:
        shop.status = "active"
        shop.plan = plan_id
        shop.plan_started_at = shop.plan_started_at or datetime.utcnow()
        shop.max_products = plan.get("max_products", shop.max_products)
        shop.max_categories = plan.get("max_categories", shop.max_categories)
        shop.plan_price = float(order.total or 0)
        base = shop.expires_at or datetime.utcnow()
        if base < datetime.utcnow():
            base = datetime.utcnow()
        shop.expires_at = base + timedelta(days=int(plan.get("days") or 30))
        db.commit()
        return {"ok": True, "verified": True, "shop": shop.to_dict(include_private=True)}

    result = aba_service.verify_payment(order, platform, transaction_id=data.transaction_id)
    if result.get("verified"):
        order.payment_status = "paid"
        order.paid_at = datetime.utcnow()
        shop.status = "active"
        shop.plan = plan_id
        shop.plan_started_at = shop.plan_started_at or datetime.utcnow()
        shop.max_products = plan.get("max_products", shop.max_products)
        shop.max_categories = plan.get("max_categories", shop.max_categories)
        shop.plan_price = float(order.total or plan.get("price", 0))
        base = shop.expires_at or datetime.utcnow()
        if base < datetime.utcnow():
            base = datetime.utcnow()
        shop.expires_at = base + timedelta(days=int(plan.get("days", 30)))
        db.commit()
        return {"ok": True, "verified": True, "shop": shop.to_dict(include_private=True)}
    return {"ok": False, "verified": False, "status": result.get("status", "pending")}


@router.post("/upgrade")
def upgrade_shop_plan(data: schemas.PlanUpgrade, db: Session = Depends(get_db),
                      user: models.User = Depends(get_current_user)):
    """Shop owner upgrades their shop's plan (1 month free / 6 months / 1 year).

    Free plans apply immediately; paid plans create an ABA (KHQR) order which is
    confirmed via /api/plans/confirm (extends the shop's expiry by the plan days).
    """
    plan_id = (data.plan or "starter").strip().lower()
    plan = PLANS.get(plan_id)
    if not plan:
        raise HTTPException(status_code=400, detail="Invalid plan")
    shop = db.query(models.Shop).filter(models.Shop.id == data.shop_id).first()
    if not shop:
        raise HTTPException(status_code=404, detail="Shop not found")
    require_shop_access(data.shop_id, user)

    free_plan = bool(plan.get("free"))
    amount = 0.0 if free_plan else round(float(plan["price"]), 2)
    now = datetime.utcnow()

    payment = None
    order_id = None
    platform = db.query(models.Shop).filter(models.Shop.username == PLATFORM_SHOP_USERNAME).first()
    if platform:
        order = models.Order(
            shop_id=platform.id,
            order_number=_gen_order_number(prefix="UPG"),
            customer_name=shop.shop_name or shop.username,
            customer_phone=shop.username,
            customer_note=f"PLAN_UPGRADE:{plan_id}",
            items_total=amount,
            total=amount,
            currency=shop.currency or "USD",
            payment_method="free" if free_plan else "khqr",
            payment_status="paid" if free_plan else "pending",
            paid_at=now if free_plan else None,
            order_status="active" if free_plan else "pending",
        )
        db.add(order)
        db.flush()
        db.add(models.OrderItem(
            order_id=order.id,
            product_name=f"Plan {plan['name']} ({plan['days']} days) — shop upgrade"
                         + (" (FREE)" if free_plan else ""),
            price=amount,
            quantity=1,
        ))
        if not free_plan:
            try:
                payment = aba_service.build_checkout_url(order, platform, success_url=data.success_url)
                order.transaction_id = payment.get("transaction_id", "")
            except PaymentNotConfigured:
                payment = None
        order_id = order.id

    if free_plan:
        # Apply immediately: extend the shop by the free plan duration.
        shop.status = "active"
        shop.plan = plan_id
        shop.plan_price = 0.0
        shop.plan_started_at = shop.plan_started_at or now
        shop.max_products = plan.get("max_products", shop.max_products)
        shop.max_categories = plan.get("max_categories", shop.max_categories)
        base = shop.expires_at or now
        if base < now:
            base = now
        shop.expires_at = base + timedelta(days=int(plan["days"]))

    db.commit()
    log_activity(db, "plan_upgrade",
                 f"Shop {shop.username} upgraded to plan {plan['name']}"
                 + (" (FREE 1 month)" if free_plan else f" (${amount})"),
                 shop.id, user)
    db.commit()

    return {
        "order_id": order_id,
        "shop_id": shop.id,
        "plan": {"id": plan_id, **plan},
        "amount": amount,
        "payment": payment,
        "free": bool(free_plan),
        "applied": bool(free_plan),
        "offer_ends": plan.get("offer_ends"),
        "expires_at": shop.expires_at.isoformat() + "Z" if shop.expires_at else None,
    }


@router.post("/resellers")
def create_reseller(data: schemas.ResellerCreate, db: Session = Depends(get_db),
                    admin: models.User = Depends(get_current_admin)):
    """Admin creates a reseller account with a referral link."""
    username = (data.username or "").strip().lower()
    if not username or not data.password:
        raise HTTPException(status_code=400, detail="Username and password are required")
    if db.query(models.User).filter(models.User.username == username).first():
        raise HTTPException(status_code=400, detail="Username already exists")
    ref = (data.referral_code or "").strip().upper() or username.upper()[:8]
    if db.query(models.User).filter(models.User.referral_code == ref).first():
        raise HTTPException(status_code=400, detail="Referral code already used")

    reseller = models.User(
        username=username,
        email=data.email,
        password_hash=hash_password(data.password),
        role="reseller",
        referral_code=ref,
        commission_rate=float(data.commission_rate or 10),
    )
    db.add(reseller)
    db.commit()
    log_activity(db, "create_reseller", f"Admin created reseller {username}", None, admin)
    db.commit()
    return {
        "id": reseller.id,
        "username": reseller.username,
        "email": reseller.email,
        "referral_code": reseller.referral_code,
        "commission_rate": reseller.commission_rate,
        "signup_link": f"{STORE_URL}/create-shop?ref={reseller.referral_code}",
    }


def _reseller_customer_data(db, shop_ids):
    """Registered customers + order-based customers (guests) across the reseller's shops."""
    registered = []
    if shop_ids:
        registered = db.query(models.Customer).filter(models.Customer.shop_id.in_(shop_ids)).all()

    order_customers = {}
    if shop_ids:
        for o in db.query(models.Order).filter(models.Order.shop_id.in_(shop_ids)).all():
            key = (o.customer_phone or o.customer_name or "").strip()
            if not key:
                continue
            if o.customer_id:
                continue  # already counted as a registered customer
            entry = order_customers.get(key)
            if entry is None:
                order_customers[key] = {
                    "name": o.customer_name or "",
                    "phone": o.customer_phone or "",
                    "shop_id": o.shop_id,
                    "orders": 0,
                }
            order_customers[key]["orders"] += 1

    rows = []
    for c in registered:
        shop = db.query(models.Shop).filter(models.Shop.id == c.shop_id).first()
        rows.append({
            "id": c.id,
            "type": "registered",
            "shop_id": c.shop_id,
            "shop_username": shop.username if shop else None,
            "name": c.name or c.username or "",
            "phone": c.phone or "",
            "telegram": c.telegram or c.telegram_username or "",
            "email": c.email or "",
            "orders": db.query(models.Order).filter(models.Order.customer_id == c.id).count(),
            "created_at": c.created_at.isoformat() if c.created_at else None,
        })
    for key, e in order_customers.items():
        shop = db.query(models.Shop).filter(models.Shop.id == e["shop_id"]).first()
        rows.append({
            "id": None,
            "type": "guest",
            "shop_id": e["shop_id"],
            "shop_username": shop.username if shop else None,
            "name": e["name"],
            "phone": e["phone"],
            "telegram": "",
            "email": "",
            "orders": e["orders"],
            "created_at": None,
        })
    return rows


@router.get("/resellers")
def list_resellers(db: Session = Depends(get_db), admin: models.User = Depends(get_current_admin)):
    """Admin: resellers with their referred shops, customers, sales and commission."""
    out = []
    for r in db.query(models.User).filter(models.User.role == "reseller").all():
        shops = db.query(models.Shop).filter(models.Shop.reseller_id == r.id).all()
        shop_ids = [s.id for s in shops]
        sales = sum(float(s.plan_price or 0) for s in shops if s.status == "active")
        commission = round(sales * float(r.commission_rate or 0) / 100.0, 2)
        customer_count = len(_reseller_customer_data(db, shop_ids))
        out.append({
            "id": r.id,
            "username": r.username,
            "email": r.email,
            "referral_code": r.referral_code,
            "commission_rate": r.commission_rate,
            "discount_max": r.discount_max if r.discount_max is not None else 1.0,
            "promo_discount": r.promo_discount if r.promo_discount is not None else 0,
            "commission_paid": r.commission_paid if r.commission_paid is not None else "not_yet",
            "commission_paid_at": r.commission_paid_at.isoformat() if r.commission_paid_at else None,
            "status": r.status,
            "shop_count": len(shops),
            "customer_count": customer_count,
            "sales": round(sales, 2),
            "commission": commission,
            "shops": [{"id": s.id, "username": s.username, "shop_name": s.shop_name,
                       "plan": s.plan, "plan_price": s.plan_price, "status": s.status}
                      for s in shops],
        })
    return out


@router.get("/resellers/{reseller_id}/customers")
def reseller_customers(reseller_id: int, db: Session = Depends(get_db),
                       admin: models.User = Depends(get_current_admin)):
    """Admin: every customer who used this reseller's code (customers of their shops)."""
    reseller = db.query(models.User).filter(models.User.id == reseller_id,
                                            models.User.role == "reseller").first()
    if not reseller:
        raise HTTPException(status_code=404, detail="Reseller not found")
    shop_ids = [s.id for s in db.query(models.Shop).filter(models.Shop.reseller_id == reseller_id).all()]
    customers = _reseller_customer_data(db, shop_ids)
    return {"reseller_id": reseller_id, "count": len(customers), "customers": customers}


@router.get("/resellers/{reseller_id}")
def get_reseller_detail(reseller_id: int, db: Session = Depends(get_db),
                        admin: models.User = Depends(get_current_admin)):
    """Admin: full detail of one reseller — profile, every referred shop (with
    counts + revenue) and all customers (registered + guest)."""
    reseller = db.query(models.User).filter(models.User.id == reseller_id,
                                            models.User.role == "reseller").first()
    if not reseller:
        raise HTTPException(status_code=404, detail="Reseller not found")
    shops = db.query(models.Shop).filter(models.Shop.reseller_id == reseller_id).all()
    shop_list = []
    for s in shops:
        paid = db.query(models.Order).filter(models.Order.shop_id == s.id,
                                             models.Order.payment_status == "paid").all()
        revenue = round(sum(o.total for o in paid), 2)
        shop_list.append({
            "id": s.id,
            "username": s.username,
            "shop_name": s.shop_name,
            "logo": s.logo,
            "plan": s.plan,
            "plan_price": s.plan_price,
            "plan_discount": s.plan_discount,
            "status": s.status,
            "currency": s.currency,
            "created_at": s.created_at.isoformat() if s.created_at else None,
            "expires_at": s.expires_at.isoformat() if s.expires_at else None,
            "product_count": db.query(models.Product).filter(models.Product.shop_id == s.id).count(),
            "category_count": db.query(models.Category).filter(models.Category.shop_id == s.id).count(),
            "order_count": db.query(models.Order).filter(models.Order.shop_id == s.id).count(),
            "customer_count": db.query(models.Customer).filter(models.Customer.shop_id == s.id).count(),
            "revenue": revenue,
        })
    shop_ids = [s.id for s in shops]
    customers = _reseller_customer_data(db, shop_ids)
    active = [s for s in shop_list if s["status"] == "active"]
    sales = sum(float(s["plan_price"] or 0) for s in active)
    commission = round(sales * float(reseller.commission_rate or 0) / 100.0, 2)
    return {
        "id": reseller.id,
        "username": reseller.username,
        "email": reseller.email,
        "referral_code": reseller.referral_code,
        "commission_rate": reseller.commission_rate,
        "discount_max": reseller.discount_max if reseller.discount_max is not None else 1.0,
        "promo_discount": reseller.promo_discount if reseller.promo_discount is not None else 0,
        "commission_paid": reseller.commission_paid if reseller.commission_paid is not None else "not_yet",
        "commission_paid_at": reseller.commission_paid_at.isoformat() if reseller.commission_paid_at else None,
        "status": reseller.status,
        "created_at": reseller.created_at.isoformat() if reseller.created_at else None,
        "signup_link": f"{STORE_URL}/create-shop?ref={reseller.referral_code}",
        "shop_count": len(shop_list),
        "customer_count": len(customers),
        "sales": round(sales, 2),
        "commission": commission,
        "shops": shop_list,
        "customers": customers,
    }


@router.put("/resellers/{reseller_id}")
def update_reseller(reseller_id: int, data: schemas.ResellerUpdate, db: Session = Depends(get_db),
                    admin: models.User = Depends(get_current_admin)):
    """Admin edits a reseller (commission, discount cap, code, status, password)."""
    reseller = db.query(models.User).filter(models.User.id == reseller_id,
                                            models.User.role == "reseller").first()
    if not reseller:
        raise HTTPException(status_code=404, detail="Reseller not found")
    if data.email is not None:
        reseller.email = data.email
    if data.password:
        reseller.password_hash = hash_password(data.password)
    if data.referral_code is not None:
        ref = (data.referral_code or "").strip().upper()
        if ref and db.query(models.User).filter(models.User.referral_code == ref,
                                                models.User.id != reseller.id).first():
            raise HTTPException(status_code=400, detail="Referral code already used")
        reseller.referral_code = ref
    if data.commission_rate is not None:
        reseller.commission_rate = float(data.commission_rate)
    if data.discount_max is not None:
        reseller.discount_max = float(data.discount_max)
    if data.status is not None:
        reseller.status = data.status
    if data.commission_paid is not None:
        paid = (data.commission_paid or "").strip().lower()
        if paid not in ("not_yet", "paid"):
            raise HTTPException(status_code=400, detail="commission_paid must be 'not_yet' or 'paid'")
        reseller.commission_paid = paid
        reseller.commission_paid_at = datetime.utcnow() if paid == "paid" else None
    db.commit()
    log_activity(db, "update_reseller", f"Admin updated reseller {reseller.username}", None, admin)
    db.commit()
    return {"ok": True, "id": reseller.id, "username": reseller.username,
            "commission_rate": reseller.commission_rate,
            "discount_max": reseller.discount_max,
            "referral_code": reseller.referral_code,
            "status": reseller.status}


@router.delete("/resellers/{reseller_id}")
def delete_reseller(reseller_id: int, db: Session = Depends(get_db),
                    admin: models.User = Depends(get_current_admin)):
    """Admin deletes a reseller; their referred shops stay but are unlinked."""
    reseller = db.query(models.User).filter(models.User.id == reseller_id,
                                            models.User.role == "reseller").first()
    if not reseller:
        raise HTTPException(status_code=404, detail="Reseller not found")
    db.query(models.Shop).filter(models.Shop.reseller_id == reseller_id).update(
        {models.Shop.reseller_id: None})
    db.delete(reseller)
    db.commit()
    log_activity(db, "delete_reseller", f"Admin deleted reseller {reseller.username}", None, admin)
    db.commit()
    return {"ok": True, "detail": "Reseller deleted"}


def _require_reseller(user: models.User):
    if user.role != "reseller":
        raise HTTPException(status_code=403, detail="Reseller access only")
    return user


def _reseller_shop_rows(db, reseller_id: int):
    shops = db.query(models.Shop).filter(models.Shop.reseller_id == reseller_id).all()
    rows = []
    for s in shops:
        rows.append({
            "id": s.id, "username": s.username, "shop_name": s.shop_name,
            "plan": s.plan, "plan_price": s.plan_price, "plan_discount": s.plan_discount,
            "status": s.status, "created_at": s.created_at.isoformat() if s.created_at else None,
            "expires_at": s.expires_at.isoformat() if s.expires_at else None,
        })
    return rows


@router.get("/reseller/me")
def reseller_me(db: Session = Depends(get_db), user: models.User = Depends(get_current_user)):
    """Reseller: own profile, referred shops, sales and commission + chart data."""
    _require_reseller(user)
    shops = _reseller_shop_rows(db, user.id)
    active = [s for s in shops if s["status"] == "active"]
    sales = sum(float(s["plan_price"] or 0) for s in active)
    commission = round(sales * float(user.commission_rate or 0) / 100.0, 2)

    # Chart data
    rate = float(user.commission_rate or 0) / 100.0
    signups = {}
    for s in shops:
        m = (s.get("created_at") or datetime.utcnow().strftime("%Y-%m-%d"))[:7]
        signups[m] = signups.get(m, 0) + 1
    months = sorted(signups.keys())[-6:]
    if not months:
        months = [datetime.utcnow().strftime("%Y-%m")]
    shops_by_status = {}
    shops_by_plan = {}
    commission_by_shop = []
    for s in shops:
        shops_by_status[s["status"]] = shops_by_status.get(s["status"], 0) + 1
        plan = s["plan"] or "none"
        shops_by_plan[plan] = shops_by_plan.get(plan, 0) + 1
        if s["status"] == "active":
            commission_by_shop.append({
                "username": s["username"], "shop_name": s["shop_name"],
                "paid": s["plan_price"], "commission": round(float(s["plan_price"] or 0) * rate, 2),
            })
    commission_by_shop.sort(key=lambda x: x["commission"], reverse=True)

    return {
        "username": user.username,
        "email": user.email,
        "referral_code": user.referral_code,
        "commission_rate": user.commission_rate,
        "discount_max": user.discount_max if user.discount_max is not None else 1.0,
        "promo_discount": user.promo_discount if user.promo_discount is not None else 0,
        "commission_paid": user.commission_paid if user.commission_paid is not None else "not_yet",
        "commission_paid_at": user.commission_paid_at.isoformat() if user.commission_paid_at else None,
        "shop_count": len(shops),
        "sales": round(sales, 2),
        "commission": commission,
        "plans": [{"id": k, **v} for k, v in PLANS.items()],
        "shops": shops,
        "charts": {
            "signup_months": months,
            "signup_series": [signups.get(m, 0) for m in months],
            "shops_by_status": shops_by_status,
            "shops_by_plan": shops_by_plan,
            "commission_by_shop": commission_by_shop[:8],
        },
    }


@router.post("/reseller/promo")
def reseller_set_promo(data: schemas.ResellerPromo, db: Session = Depends(get_db),
                       user: models.User = Depends(get_current_user)):
    """Reseller saves their default promo discount (auto-applied with their code)."""
    _require_reseller(user)
    cap = float(user.discount_max if user.discount_max is not None else 1.0)
    promo = max(0.0, min(float(data.promo_discount or 0), cap))
    user.promo_discount = round(promo, 2)
    db.commit()
    return {"ok": True, "promo_discount": user.promo_discount, "discount_max": cap}


@router.get("/reseller/export")
def reseller_export(export_format: str = "zip", db: Session = Depends(get_db),
                    user: models.User = Depends(get_current_user)):
    """Reseller: download a full backup (JSON/ZIP) or Excel report of their data."""
    _require_reseller(user)
    fmt = (export_format or "zip").lower()
    shops = db.query(models.Shop).filter(models.Shop.reseller_id == user.id).all()
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    os.makedirs(app_config.BACKUP_DIR, exist_ok=True)

    if fmt == "xlsx":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reseller"
        ws.append(["Username", user.username])
        ws.append(["Referral Code", user.referral_code])
        ws.append(["Commission Rate %", user.commission_rate])
        ws.append(["Max Discount $", user.discount_max])
        ws2 = wb.create_sheet("Shops")
        ws2.append(["ID", "Username", "Shop Name", "Plan", "Plan Price $", "Discount $", "Status", "Expires"])
        for s in shops:
            ws2.append([s.id, s.username, s.shop_name, s.plan, s.plan_price,
                        s.plan_discount, s.status,
                        s.expires_at.strftime("%Y-%m-%d") if s.expires_at else ""])
        ws3 = wb.create_sheet("Commissions")
        ws3.append(["Shop", "Plan Price $", "Commission Rate %", "Commission $"])
        for s in shops:
            ws3.append([s.username, s.plan_price, user.commission_rate,
                        round(float(s.plan_price or 0) * float(user.commission_rate or 0) / 100.0, 2)])
        path = os.path.join(app_config.BACKUP_DIR, f"reseller_{user.username}_{ts}.xlsx")
        wb.save(path)
        return FileResponse(path, filename=os.path.basename(path),
                            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if fmt == "json":
        data = {
            "version": 1, "type": "reseller",
            "exported_at": datetime.utcnow().isoformat(),
            "reseller": {"username": user.username, "email": user.email,
                         "referral_code": user.referral_code,
                         "commission_rate": user.commission_rate,
                         "discount_max": user.discount_max},
            "shops": [backup_service._collect_shop_data(db, s.id) for s in shops],
        }
        path = os.path.join(app_config.BACKUP_DIR, f"reseller_{user.username}_{ts}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return FileResponse(path, filename=os.path.basename(path), media_type="application/json")

    # default: ZIP with one JSON per shop + summary
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        summary = {
            "version": 1, "type": "reseller", "exported_at": datetime.utcnow().isoformat(),
            "reseller": {"username": user.username, "email": user.email,
                         "referral_code": user.referral_code,
                         "commission_rate": user.commission_rate,
                         "discount_max": user.discount_max},
            "shop_count": len(shops),
        }
        zf.writestr("reseller_info.json", json.dumps(summary, ensure_ascii=False, indent=2))
        for s in shops:
            zf.writestr(f"shop_{s.username}.json",
                        json.dumps(backup_service._collect_shop_data(db, s.id),
                                   ensure_ascii=False, indent=2))
    path = os.path.join(app_config.BACKUP_DIR, f"reseller_{user.username}_{ts}.zip")
    with open(path, "wb") as f:
        f.write(buf.getvalue())
    return FileResponse(path, filename=os.path.basename(path), media_type="application/zip")


@router.post("/reseller/import")
def reseller_import(file: UploadFile = File(...), db: Session = Depends(get_db),
                    user: models.User = Depends(get_current_user)):
    """Reseller: restore data from their own JSON/ZIP backup into their shops."""
    _require_reseller(user)
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in (".json", ".zip"):
        raise HTTPException(status_code=400, detail="Only .json or .zip backups can be imported")
    data = file.file.read()
    tmp = tempfile.mkdtemp(prefix="reseller_import_")
    saved = os.path.join(tmp, f"import{ext}")
    with open(saved, "wb") as f:
        f.write(data)

    count, skipped = 0, 0
    try:
        if ext == ".zip":
            with zipfile.ZipFile(saved, "r") as zf:
                for name in zf.namelist():
                    if not name.endswith(".json"):
                        continue
                    inner = os.path.join(tmp, os.path.basename(name))
                    with open(inner, "wb") as f:
                        f.write(zf.read(name))
                    try:
                        c, sk = _import_one_reseller_shop(db, inner, user)
                        count += c; skipped += sk
                    except Exception:
                        skipped += 1
        else:
            try:
                c, sk = _import_one_reseller_shop(db, saved, user)
                count += c; skipped += sk
            except Exception:
                skipped += 1
    finally:
        import shutil
        shutil.rmtree(tmp, ignore_errors=True)
    db.commit()
    return {"ok": True, "records_restored": count, "duplicates_skipped": skipped}


def _import_one_reseller_shop(db, filepath: str, user: models.User):
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    # reseller bundle -> restore each shop payload; single shop backup -> one payload
    payloads = data.get("shops") or [data] if data.get("type") == "reseller" else [data]
    total, skipped, total_images = 0, 0, 0
    for payload in payloads:
        shop_dict = (payload or {}).get("shop") or {}
        username = (shop_dict.get("username") or "").strip()
        if not username:
            skipped += 1
            continue
        shop = db.query(models.Shop).filter(models.Shop.username == username).first()
        if not shop or shop.reseller_id != user.id:
            skipped += 1
            continue
        # import_shop_backup expects a single-shop backup file — write one per payload
        tmp = tempfile.NamedTemporaryFile("w", suffix=".json", delete=False, encoding="utf-8")
        json.dump(payload, tmp, ensure_ascii=False)
        tmp.close()
        try:
            c, sk, imgs = backup_service.import_shop_backup(db, shop.id, tmp.name)
            total += c; skipped += sk; total_images += imgs
        finally:
            os.remove(tmp.name)
    return total, skipped




@router.get("/charts")
def platform_charts(db: Session = Depends(get_db), admin: models.User = Depends(get_current_admin)):
    """Admin: real platform-wide data for dashboard charts."""
    from datetime import timedelta as _td

    # 1) Revenue by day (last 30 days)
    since = datetime.utcnow() - _td(days=29)
    by_day = {}
    for o in db.query(models.Order).filter(models.Order.payment_status == "paid",
                                           models.Order.paid_at >= since).all():
        day = (o.paid_at or o.created_at).strftime("%Y-%m-%d")
        by_day[day] = round(by_day.get(day, 0) + (o.total or 0), 2)
    days, revenue_series = [], []
    for i in range(30):
        d = (since + _td(days=i)).strftime("%Y-%m-%d")
        days.append(d)
        revenue_series.append(by_day.get(d, 0))

    # 2) Orders by payment status
    orders_by_status = [{"status": k or "unknown", "count": v}
                        for k, v in db.query(models.Order.payment_status,
                                             func.count(models.Order.id))
                        .group_by(models.Order.payment_status).all()]

    # 3) Shops by plan
    shops_by_plan = [{"plan": k or "none", "count": v}
                     for k, v in db.query(models.Shop.plan, func.count(models.Shop.id))
                     .group_by(models.Shop.plan).all()]

    # 4) Top shops by paid revenue
    top = []
    for s in db.query(models.Shop).all():
        rev = db.query(func.coalesce(func.sum(models.Order.total), 0)).filter(
            models.Order.shop_id == s.id, models.Order.payment_status == "paid").scalar() or 0
        top.append({"id": s.id, "username": s.username, "shop_name": s.shop_name,
                    "revenue": round(float(rev), 2)})
    top.sort(key=lambda x: x["revenue"], reverse=True)

    # 5) Resellers: commission + payment status
    resellers = []
    for r in db.query(models.User).filter(models.User.role == "reseller").all():
        rshops = db.query(models.Shop).filter(models.Shop.reseller_id == r.id,
                                              models.Shop.status == "active").all()
        sales = sum(float(s.plan_price or 0) for s in rshops)
        resellers.append({
            "username": r.username,
            "sales": round(sales, 2),
            "commission": round(sales * float(r.commission_rate or 0) / 100.0, 2),
            "commission_paid": r.commission_paid or "not_yet",
        })

    # 6) Shop signups by month (last 6 months)
    signups = {}
    for s in db.query(models.Shop).all():
        m = (s.created_at or datetime.utcnow()).strftime("%Y-%m")
        signups[m] = signups.get(m, 0) + 1
    months = sorted(signups.keys())[-6:]
    if not months:
        months = [datetime.utcnow().strftime("%Y-%m")]

    return {
        "revenue_days": days,
        "revenue_series": revenue_series,
        "orders_by_status": orders_by_status,
        "shops_by_plan": shops_by_plan,
        "top_shops": top[:6],
        "resellers": resellers,
        "signup_months": months,
        "signup_series": [signups.get(m, 0) for m in months],
    }
